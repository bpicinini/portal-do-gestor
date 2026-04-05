"""Script para popular a base inicial com dados reais extraídos das planilhas."""
import os
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

from utils.excel_io import (
    carregar_workbook, salvar_workbook,
    SHEET_DEPARTAMENTOS, SHEET_CARGOS, SHEET_COLABORADORES,
    SHEET_HISTORICO, SHEET_MANPOWER_MENSAL, SHEET_PERFORMANCE,
)

# ── IDs novos por data de admissão ────────────────────────────────────────────
# Critério de desempate: departamento_id → cargo_nivel → nome alfabético
# 2020-01-01 legacy → IDs 1–47, depois entradas recentes em ordem cronológica

def popular():
    wb = carregar_workbook()

    def unidade_por_cidade(cidade):
        return "Itajaí" if cidade == "Itajaí" else "Novo Hamburgo"

    # === DEPARTAMENTOS ===
    ws = wb[SHEET_DEPARTAMENTOS]
    departamentos = [
        (1, "Importação"),
        (2, "Agenciamento"),
        (3, "Exportação"),
        (4, "Seguro Internacional"),
    ]
    for dept_id, nome in departamentos:
        ws.append([dept_id, nome])

    # === CARGOS ===
    # id, nome, nivel, peso_manpower, departamento_id
    ws = wb[SHEET_CARGOS]
    cargos = [
        # Gestão Geral
        (99, "Gerente de Operações",    1,   None, 1),
        # Importação
        (1,  "Coordenador",             2,   None, 1),
        (2,  "Supervisor",              2.5, None, 1),
        (3,  "Especialista",            3,   1.5,  1),
        (4,  "Analista Sênior",         4,   1.3,  1),
        (5,  "Analista Pleno",          5,   1.15, 1),
        (6,  "Analista Júnior",         6,   1.0,  1),
        (7,  "Assistente",              7,   0.5,  1),
        (8,  "Estagiário",              8,   0.25, 1),
        (9,  "Jovem Aprendiz",          9,   0.25, 1),
        (14, "Assistente",              7,   None, 1),
        # Agenciamento
        (20, "Coordenador",             2,   None, 2),
        (21, "Supervisor",              2.5, None, 2),
        (22, "Analista Sênior",         4,   1.3, 2),
        (23, "Assistente",              7,   0.5, 2),
        (24, "Assistente",              7,   0.5, 2),
        (25, "Assistente",              7,   0.5, 2),
        (26, "Analista Júnior",         6,   1.0, 2),
        # Exportação
        (30, "Supervisor",              2.5, None, 3),
        (31, "Especialista",            3,   1.5,  3),
        (32, "Analista Sênior",         4,   1.3,  3),
        (33, "Analista Pleno",          5,   1.15, 3),
        (34, "Assistente",              7,   0.5,  3),
        # Seguro Internacional — agora com peso manpower
        (40, "Analista Pleno",          5,   1.15, 4),
        (41, "Jovem Aprendiz",          9,   0.25, 4),
    ]
    for row in cargos:
        ws.append(list(row))

    # === COLABORADORES ===
    # id, nome, cargo_id, departamento_id, empresa, cidade,
    # tipo_contrato, gestor_direto, data_entrada, data_saida, status
    ws = wb[SHEET_COLABORADORES]
    colaboradores = [
        # ── Gerência Geral ─────────────────────────────────────────────────────
        (1,  "Bruno Picinini",                        99, 1, "3S Corp",            "Novo Hamburgo", "", "Gabriel Spohr",   date(2020,1,1),  None,            "Ativo"),
        # ── Importação — Coordenação ───────────────────────────────────────────
        (2,  "Patrice Zeidler Basso",                 1,  1, "Leader NH",          "Novo Hamburgo", "", "Bruno Picinini",   date(2020,1,1),  None,            "Ativo"),
        (3,  "Nilson Tafarel de Quadros",             1,  1, "Leader SC",          "Itajaí",        "", "Bruno Picinini",   date(2020,1,1),  date(2025,12,1), "Inativo"),
        # ── Importação — Supervisão ────────────────────────────────────────────
        (4,  "Mariana Strick",                        2,  1, "Leader SC",          "Itajaí",        "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        # ── Importação — Especialista ──────────────────────────────────────────
        (5,  "Adir Gomes Carvalho",                   3,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        # ── Importação — Analistas Sênior ─────────────────────────────────────
        (6,  "Bruna Sandy Dionizio Schell",           4,  1, "Leader SC",          "Itajaí",        "", "Mariana Strick",   date(2020,1,1),  None,            "Ativo"),
        (7,  "Lucimar Barbosa Serpa",                 4,  1, "Winning Trading ITJ","Itajaí",        "", "Mariana Strick",   date(2020,1,1),  None,            "Ativo"),
        (8,  "Marcelo Kirsch",                        4,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (9,  "Regina dos Santos Cardoso",             4,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (10, "Ricardo Guimarães Passos Costa",        4,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        # ── Importação — Analistas Pleno ──────────────────────────────────────
        (11, "Caroline Florêncio de Andrade",         5,  1, "Winning Trading ITJ","Itajaí",        "", "Mariana Strick",   date(2020,1,1),  None,            "Ativo"),
        (12, "Fernanda Farias Dorneles de Carvalho",  5,  1, "Leader SC",          "Itajaí",        "", "Mariana Strick",   date(2020,1,1),  None,            "Ativo"),
        (13, "Franciely Meirelles Artigas",           5,  1, "Winning Trading ITJ","Itajaí",        "", "Mariana Strick",   date(2020,1,1),  None,            "Ativo"),
        (14, "Renan Gomes Dias",                      5,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (15, "Tainá da Silva Ripcke",                 5,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        # ── Importação — Analistas Júnior ─────────────────────────────────────
        (16, "Camilli Diovana da Silveira",           6,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (17, "Giovanna Vitória Sebben",               6,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (18, "Nicoly Cordeiro Lins",                  6,  1, "Leader SC",          "Itajaí",        "", "Mariana Strick",   date(2020,1,1),  None,            "Ativo"),
        # ── Importação — Assistentes ──────────────────────────────────────────
        (19, "Ana Luiza Silva",                       7,  1, "Winning Trading ITJ","Itajaí",        "", "Mariana Strick",   date(2020,1,1),  None,            "Ativo"),
        (20, "Juliano Kairê Sperandei",               14, 1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (21, "Lívia Moehlecke Haubert",               7,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (22, "Lucas Leite Silva da Rosa",             7,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (23, "Vitória Ayres de Souza",                7,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (24, "Vitor Matheus Duarte",                  7,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        # ── Importação — Inativos históricos ──────────────────────────────────
        (25, "Ramil Acosta",                          5,  1, "Winning Trading ITJ","Itajaí",        "", "Mariana Strick",   date(2020,1,1),  date(2026,1,1),  "Inativo"),
        (26, "Beatriz Silva",                         7,  1, "Leader SC",          "Itajaí",        "", "Mariana Strick",   date(2020,1,1),  date(2026,3,1),  "Inativo"),
        # ── Agenciamento ──────────────────────────────────────────────────────
        (27, "Gabriel da Silva Hilgert",              20, 2, "Leader NH",          "Novo Hamburgo", "", "Bruno Picinini",             date(2021,2,8),  None, "Ativo"),
        (28, "Matheus Lara Ropelato",                 21, 2, "Leader SC",          "Novo Hamburgo", "", "Gabriel da Silva Hilgert",   date(2022,3,7),  None, "Ativo"),
        (29, "Tiago Schneider",                       22, 2, "Leader NH",          "Novo Hamburgo", "", "Gabriel da Silva Hilgert",   date(2021,11,16),None, "Ativo"),
        (30, "Agata Trinity Gonçalves Machado",       23, 2, "Leader NH",          "Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2024,6,3),  None, "Ativo"),
        (31, "Augusto da Luz Schneider",              24, 2, "Leader NH",          "Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2025,12,17),None, "Ativo"),
        (32, "Beatriz da Silva",                      24, 2, "Winning Trading ITJ","Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2024,2,5),  None, "Ativo"),
        (33, "Franciane de Oliveira",                 24, 2, "Leader SC",          "Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2024,12,9), None, "Ativo"),
        (34, "Graziele Marques da Silva Meytre",      25, 2, "Leader SC",          "Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2023,12,4), None, "Ativo"),
        (35, "Kailane Eduarda Claudino de Moura",     23, 2, "Leader NH",          "Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2023,3,8),  None, "Ativo"),
        (36, "Laiane Rosiane da Luz",                 23, 2, "Leader SC",          "Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2023,5,8),  None, "Ativo"),
        (37, "Manuella Gonçalves da Silva",           23, 2, "Leader SC",          "Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2023,12,4), None, "Ativo"),
        (38, "Mariana Matte Reis",                    23, 2, "Leader NH",          "Novo Hamburgo", "", "Matheus Lara Ropelato",      date(2021,11,8), None, "Ativo"),
        # ── Exportação ────────────────────────────────────────────────────────
        (39, "Liliane Maus",                          30, 3, "3S Corp",            "Novo Hamburgo", "", "Bruno Picinini",   date(2020,1,1),  None,            "Ativo"),
        (40, "Felipe Marcellino",                     31, 3, "3S Corp",            "Novo Hamburgo", "", "Liliane Maus",     date(2020,1,1),  None,            "Ativo"),
        (41, "Elizandra Severo Moraes",               32, 3, "3S Corp",            "Novo Hamburgo", "", "Liliane Maus",     date(2020,1,1),  None,            "Ativo"),
        (42, "Eliane Teresa Lemes Bielefeld",         33, 3, "3S Corp",            "Novo Hamburgo", "", "Liliane Maus",     date(2020,1,1),  None,            "Ativo"),
        (43, "Jéssica Machado Oliveira",              33, 3, "3S Corp",            "Novo Hamburgo", "", "Liliane Maus",     date(2020,1,1),  None,            "Ativo"),
        (44, "Roberta da Rosa Melo",                  33, 3, "3S Corp",            "Novo Hamburgo", "", "Liliane Maus",     date(2020,1,1),  None,            "Ativo"),
        (45, "Sumaia Gabriele Schneider",             34, 3, "3S Corp",            "Novo Hamburgo", "", "Liliane Maus",     date(2020,1,1),  None,            "Ativo"),
        # ── Seguro Internacional ──────────────────────────────────────────────
        (46, "Camila de Souza Nunes",                 40, 4, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        (47, "Isadora Valentina Soares",              41, 4, "3S Corp",            "Novo Hamburgo", "", "Patrice Basso",    date(2020,1,1),  None,            "Ativo"),
        # ── Entradas recentes (ordem cronológica) ─────────────────────────────
        (48, "Eduarda Silva",                         7,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2025,8,1),  None,            "Ativo"),
        (49, "Beatriz Vieira",                        7,  1, "Leader SC",          "Itajaí",        "", "Mariana Strick",   date(2025,9,1),  None,            "Ativo"),
        (50, "Mayara Campos",                         8,  1, "Leader SC",          "Itajaí",        "", "Patrice Basso",    date(2025,9,1),  None,            "Ativo"),
        (51, "Mateus Rauber",                         8,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2025,10,1), None,            "Ativo"),
        (52, "Giovanna Pampanelli",                   6,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2025,11,1), None,            "Ativo"),
        (53, "João Alves",                            6,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2025,11,1), None,            "Ativo"),
        (54, "Aline Grimm",                           4,  1, "Leader NH",          "Novo Hamburgo", "", "Mariana Strick",   date(2025,12,1), None,            "Ativo"),
        (55, "Lucas Perin",                           7,  1, "Leader NH",          "Novo Hamburgo", "", "Mariana Strick",   date(2026,2,1),  None,            "Ativo"),
        (56, "Nicoly Bizarro",                        8,  1, "Leader NH",          "Novo Hamburgo", "", "Patrice Basso",    date(2026,3,1),  None,            "Ativo"),
        (57, "Wemilly Carvalho",                      8,  1, "Leader NH",          "Novo Hamburgo", "", "Mariana Strick",   date(2026,3,1),  None,            "Ativo"),
        (58, "Bruna Kelsch da Silva",                 22, 2, "Leader NH",          "Novo Hamburgo", "", "Gabriel da Silva Hilgert",   date(2020,10,9), None, "Ativo"),
        (59, "Enila Kawely da Silva Padilha",         22, 2, "Leader NH",          "Novo Hamburgo", "", "Gabriel da Silva Hilgert",   date(2022,5,9),  None, "Ativo"),
        (60, "Mariana de Sousa Piccini",              26, 2, "Leader SC",          "Novo Hamburgo", "", "Gabriel da Silva Hilgert",   date(2026,3,16), None, "Ativo"),
        (61, "Matheus Bourscheid",                    26, 2, "Leader NH",          "Novo Hamburgo", "", "Gabriel da Silva Hilgert",   date(2022,4,18), None, "Ativo"),
    ]
    for row in colaboradores:
        row = list(row)
        row[4] = unidade_por_cidade(row[5])
        ws.append(row)

    # === HISTÓRICO ===
    # id, colaborador_id, nome, tipo, cargo, data, observação
    ws = wb[SHEET_HISTORICO]
    historico = [
        (1,  3,    "Nilson Tafarel de Quadros", "Saída",   "Coordenador",     date(2025,12,1), ""),
        (2,  57,   "Wemilly Carvalho",           "Entrada", "Estagiário",      date(2026,3,1),  ""),
        (3,  56,   "Nicoly Bizarro",             "Entrada", "Estagiário",      date(2026,3,1),  ""),
        (4,  26,   "Beatriz Silva",              "Saída",   "Assistente",      date(2026,3,1),  ""),
        (5,  55,   "Lucas Perin",                "Entrada", "Assistente",      date(2026,2,1),  ""),
        (6,  25,   "Ramil Acosta",               "Saída",   "Analista Pleno",  date(2026,1,1),  ""),
        (7,  54,   "Aline Grimm",                "Entrada", "Analista Sênior", date(2025,12,1), ""),
        (8,  52,   "Giovanna Pampanelli",        "Entrada", "Analista Júnior", date(2025,11,1), ""),
        (9,  53,   "João Alves",                 "Entrada", "Analista Júnior", date(2025,11,1), ""),
        (10, 51,   "Mateus Rauber",              "Entrada", "Estagiário",      date(2025,10,1), ""),
        (11, 50,   "Mayara Campos",              "Entrada", "Estagiário",      date(2025,9,1),  ""),
        (12, 49,   "Beatriz Vieira",             "Entrada", "Assistente",      date(2025,9,1),  ""),
        (13, 48,   "Eduarda Silva",              "Entrada", "Assistente",      date(2025,8,1),  ""),
        (14, None, "Bruna Romanos",              "Saída",   "Analista Júnior", date(2025,7,1),  ""),
        (15, None, "Igor Rebello",               "Saída",   "Assistente",      date(2025,6,1),  ""),
        (16, None, "Ewellen Aquino",             "Saída",   "Analista Júnior", date(2025,5,1),  ""),
        (17, None, "Giordana Allgayer",          "Saída",   "Analista Sênior", date(2025,3,1),  ""),
        (18, None, "Felipe Saueressig",          "Saída",   "Analista Sênior", date(2025,2,1),  ""),
        (19, None, "Erick Andrade",              "Saída",   "Assistente",      date(2024,12,1), ""),
        (20, None, "Arthur Krause",              "Saída",   "Assistente",      date(2024,12,1), ""),
        (21, None, "Emilin Fagundes",            "Saída",   "Analista Sênior", date(2024,9,1),  ""),
        (22, None, "Jennifer Schneider",         "Saída",   "Analista Sênior", date(2024,9,1),  ""),
        (23, None, "Julio Passos",               "Saída",   "Assistente",      date(2024,8,1),  ""),
        (24, None, "Leonardo Lima",              "Saída",   "Analista Pleno",  date(2024,8,1),  ""),
        (25, None, "Emilin Fagundes",            "Entrada", "Analista Sênior", date(2024,8,1),  ""),
    ]
    for row in historico:
        ws.append(list(row))

    # === PERFORMANCE ===
    ws = wb[SHEET_PERFORMANCE]
    meta_2025 = 335.56

    perf_2024 = [
        (2024, 7,  7004, 32.75),
        (2024, 8,  7663, 31.55),
        (2024, 9,  8832, 29.90),
        (2024, 10, 8717, 27.30),
        (2024, 11, 8682, 27.30),
        (2024, 12, 8345, 27.30),
    ]
    for ano, mes, vol, mp in perf_2024:
        perf = round(vol / mp, 2)
        ws.append([ano, mes, vol, mp, perf, None, None])

    perf_2025 = [
        (2025, 1,  9429, 26.30),
        (2025, 2,  7634, 24.80),
        (2025, 3,  8068, 23.50),
        (2025, 4,  7244, 23.50),
        (2025, 5,  7343, 23.50),
        (2025, 6,  7792, 23.00),
        (2025, 7,  8564, 22.00),
        (2025, 8,  8566, 22.65),
        (2025, 9,  9208, 22.90),
        (2025, 10, 8988, 23.65),
        (2025, 11, 8409, 25.65),
        (2025, 12, 8393, 25.45),
    ]
    for ano, mes, vol, mp in perf_2025:
        perf = round(vol / mp, 2)
        pct  = round(perf / meta_2025, 4)
        ws.append([ano, mes, vol, mp, perf, meta_2025, pct])

    meta_2026_s1 = 361.42
    perf_2026 = [
        (2026, 1, 10721, 24.30),
        (2026, 2, 10040, 24.30),
        (2026, 3, 10700, 24.30),
    ]
    for ano, mes, vol, mp in perf_2026:
        perf = round(vol / mp, 2)
        pct  = round(perf / meta_2026_s1, 4)
        ws.append([ano, mes, vol, mp, perf, meta_2026_s1, pct])

    salvar_workbook(wb)
    print("Base populada com sucesso!")
    print(f"  Departamentos: {len(departamentos)}")
    print(f"  Cargos: {len(cargos)}")
    print(f"  Colaboradores: {len(colaboradores)}")
    print(f"  Histórico: {len(historico)} registros")
    print(f"  Performance: {len(perf_2024)+len(perf_2025)+len(perf_2026)} meses")


def gerar_planilha_colaboradores():
    """Gera colaboradores.xlsx na raiz do projeto com formatação."""
    from openpyxl import Workbook as OWB
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Ler dados do banco
    wb_dados = carregar_workbook()
    from utils.excel_io import sheet_to_list, SHEET_DEPARTAMENTOS, SHEET_CARGOS
    colaboradores_raw = sheet_to_list(wb_dados[SHEET_COLABORADORES])
    cargos_map = {c["id"]: c for c in sheet_to_list(wb_dados[SHEET_CARGOS])}
    deptos_map = {d["id"]: d["nome"] for d in sheet_to_list(wb_dados[SHEET_DEPARTAMENTOS])}

    wb = OWB()
    ws = wb.active
    ws.title = "Colaboradores"

    # Cabeçalhos
    headers = ["ID", "Nome", "Departamento", "Cargo", "Unidade",
               "Gestor Direto", "Admissão", "Saída", "Status"]
    col_widths = [5, 38, 22, 25, 22, 22, 12, 12, 10]

    navy    = "1E3A5F"
    white   = "FFFFFF"
    ativo   = "E8F5E9"
    inativo = "FAFAFA"
    thin    = Side(style="thin", color="CCCCCC")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=white, size=11)
        cell.fill      = PatternFill("solid", fgColor=navy)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 22

    meses_pt = ["","Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]

    def fmt_data(d):
        if d is None:
            return ""
        try:
            from datetime import datetime
            if isinstance(d, str):
                d = datetime.strptime(d, "%Y-%m-%d")
            return f"{meses_pt[d.month]}/{d.year}"
        except Exception:
            return str(d)

    # Dados — ordenar por admissão, depois nome
    colaboradores_raw.sort(key=lambda c: (
        str(c.get("data_entrada") or ""),
        c.get("nome", "")
    ))

    for row_i, c in enumerate(colaboradores_raw, 2):
        cargo  = cargos_map.get(c.get("cargo_id"), {})
        status = c.get("status", "")
        fill_color = ativo if status == "Ativo" else inativo

        valores = [
            c.get("id"),
            c.get("nome"),
            deptos_map.get(c.get("departamento_id"), "—"),
            cargo.get("nome", "—"),
            c.get("cidade") if c.get("cidade") in ("Novo Hamburgo", "Itajaí") else c.get("empresa"),
            c.get("gestor_direto"),
            fmt_data(c.get("data_entrada")),
            fmt_data(c.get("data_saida")),
            status,
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.border    = border
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if col in (1,7,8,9) else "left")
            if col == 9:
                color = "2E7D32" if status == "Ativo" else "757575"
                cell.font = Font(bold=True, color=color)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    project_root = os.path.dirname(os.path.dirname(__file__))
    path = os.path.join(project_root, "colaboradores.xlsx")
    wb.save(path)
    print(f"  Planilha colaboradores.xlsx gerada em: {path}")


if __name__ == "__main__":
    popular()
    gerar_planilha_colaboradores()
