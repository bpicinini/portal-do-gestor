"""Script para importar dados de Processos Judiciais do Excel para o banco."""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from utils.processos_judiciais import inserir_em_lote

# Mapeamento de status do Excel para status do sistema
STATUS_MAP = {
    "ENCERRADO EM DEFINITIVO": "Encerrado",
    "ETAPA FINALIZADA": "Encerrado",
    "INTIMAÇÃO RESPONDIDA": "Intimação Respondida",
    "INTIMAÇÃO PENDENTE": "Intimação Pendente",
    "JUDICIALIZADO": "Judicializado",
    "ANÁLISE / TRIAGEM": "Análise/Triagem",
}


def importar_processos_judiciais():
    excel_file = "Processos Judiciais.xlsx"
    wb = load_workbook(excel_file)
    ws = wb.active

    registros = []
    current_status = None

    for row_idx in range(1, ws.max_row + 1):
        # Coluna A (primeira célula da linha)
        first_cell = ws.cell(row=row_idx, column=1).value

        # Detectar mudança de status
        if first_cell and first_cell in STATUS_MAP:
            current_status = STATUS_MAP[first_cell]
            print(f"Encontrado status: {first_cell} → {current_status}")
            continue

        # Detectar headers (linha com "Task Name" na coluna B)
        second_cell = ws.cell(row=row_idx, column=2).value
        if second_cell and str(second_cell).strip() == "Task Name":
            print(f"Encontrado header em row {row_idx}")
            continue

        # Pular linhas vazias ou que não têm dados
        task_name = second_cell
        if not task_name or not str(task_name).strip():
            continue

        # Extrair dados
        tipo_processo = ws.cell(row=row_idx, column=3).value
        parte_contraria = ws.cell(row=row_idx, column=4).value
        cliente = ws.cell(row=row_idx, column=5).value
        numero_processo = ws.cell(row=row_idx, column=6).value
        prazo_fatal = ws.cell(row=row_idx, column=7).value
        valor = ws.cell(row=row_idx, column=8).value
        comentario = ws.cell(row=row_idx, column=9).value
        data_maturacao = ws.cell(row=row_idx, column=10).value

        # Normalizar valores
        registro = {
            "titulo": str(task_name).strip(),
            "tipo_processo": str(tipo_processo).strip() if tipo_processo else "Administrativo",
            "parte_contraria": str(parte_contraria).strip() if parte_contraria else None,
            "cliente": str(cliente).strip() if cliente else None,
            "numero_processo": str(numero_processo).strip() if numero_processo else None,
            "prazo_fatal": prazo_fatal,
            "valor": float(valor) if valor else None,
            "data_maturacao": data_maturacao,
            "status": current_status or "Análise/Triagem",
        }

        # Adicionar comentário inicial se houver
        if comentario and str(comentario).strip():
            registro["comentarios_iniciais"] = [
                {
                    "tipo": "comentario",
                    "usuario": "sistema",
                    "data_hora": None,
                    "texto": str(comentario).strip(),
                }
            ]

        registros.append(registro)
        print(f"  Row {row_idx}: {registro['titulo'][:50]}... → {registro['status']}")

    if registros:
        print(f"\nImportando {len(registros)} processos...")
        ids = inserir_em_lote(registros, usuario="migração")
        print(f"✅ Importação concluída! {len(ids)} processos inseridos.")
        return ids
    else:
        print("❌ Nenhum registro encontrado para importar.")
        return []


if __name__ == "__main__":
    importar_processos_judiciais()
