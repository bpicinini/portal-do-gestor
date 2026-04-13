#!/usr/bin/env python3
"""
Atualiza dados.xlsx com todas as mudanças pendentes (Abril 2026):
1. Lucas Perin (id=55): cidade/empresa → Itajaí
2. Cria cargo Coordenador para Seguro Internacional (dept 4)
3. Adiciona Patrice como Coordenadora de Seguro Internacional (dept 4)
4. Ana Luísa (id=19): cargo_id 7 → 6 (Assistente → Analista Júnior)
5. Beatriz Vieira (id=49): cargo_id 7 → 8 (Assistente → Estagiário)
6. Vitória (id=23): desligamento em 2026-04-13
7. Juliano (id=20): responsavel_direto = "Patrice Zeidler Basso"
8. Recalcula Manpower e Performance de Abril/2026
"""
import os
import sys
from datetime import date, datetime
from calendar import monthrange

# Add project root to path
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

from openpyxl import load_workbook

DATA_PATH = os.path.join(_PROJECT_ROOT, "data", "dados.xlsx")


def col_idx(ws, header_name):
    """Retorna o índice (1-based) da coluna com esse header."""
    for cell in ws[1]:
        if str(cell.value or "").strip().lower() == header_name.lower():
            return cell.column
    return None


def find_row(ws, col_num, value):
    """Retorna o número da linha onde ws[row][col_num] == value (1-based col)."""
    for row in ws.iter_rows(min_row=2):
        cell = row[col_num - 1]
        if cell.value is not None and str(cell.value).strip() == str(value).strip():
            return cell.row
    return None


def max_id(ws):
    """Retorna o maior id numérico da coluna 1."""
    max_val = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            try:
                max_val = max(max_val, int(row[0]))
            except (ValueError, TypeError):
                pass
    return max_val


def status_efetivo(colaborador_dict, referencia):
    def norm_data(val):
        if val is None or val == "":
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        if isinstance(val, str):
            for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(val, fmt).date()
                except ValueError:
                    continue
        return None

    entrada = norm_data(colaborador_dict.get("data_entrada"))
    saida = norm_data(colaborador_dict.get("data_saida"))
    status = colaborador_dict.get("status")

    if entrada and entrada > referencia:
        return "Inativo"
    if saida and saida <= referencia:
        return "Inativo"
    if status == "Inativo" and saida and saida > referencia:
        return "Ativo"
    if status in ("Ativo", "Inativo"):
        return status
    return "Ativo"


def main():
    print(f"Carregando {DATA_PATH}...")
    wb = load_workbook(DATA_PATH)

    ws_colab = wb["Colaboradores"]
    ws_cargos = wb["Cargos"]
    ws_hist = wb["Historico"]
    ws_mp = wb["ManpowerMensal"]
    ws_perf = wb["Performance"]

    # ── Lê headers das sheets ──────────────────────────────────────────────
    def sheet_to_dicts(ws):
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 1:
            return [], []
        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
        records = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            records.append(dict(zip(headers, row)))
        return headers, records

    # ── 1. Lucas Perin (id=55): cidade → Itajaí, empresa → Itajaí ─────────
    print("\n[1] Atualizando Lucas Perin (id=55) → Itajaí")
    row_lucas = find_row(ws_colab, col_idx(ws_colab, "id"), 55)
    if row_lucas:
        col_empresa = col_idx(ws_colab, "empresa")
        col_cidade = col_idx(ws_colab, "cidade")
        ws_colab.cell(row=row_lucas, column=col_empresa, value="Itajaí")
        ws_colab.cell(row=row_lucas, column=col_cidade, value="Itajaí")
        print(f"  ✓ Lucas Perin linha {row_lucas}: empresa e cidade → Itajaí")
    else:
        print("  ✗ Lucas Perin (id=55) não encontrado!")

    # ── 2. Cria cargo Coordenador para Seguro Internacional (dept 4) ───────
    print("\n[2] Criando cargo Coordenador para Seguro Internacional (dept 4)")
    _, cargos = sheet_to_dicts(ws_cargos)

    # Verifica se já existe Coordenador em dept 4
    coord_dept4 = next(
        (c for c in cargos
         if str(c.get("departamento_id", "")).strip() == "4"
         and str(c.get("nivel", "")).strip() in ("2", "2.0")),
        None
    )
    if coord_dept4:
        novo_cargo_id = int(coord_dept4["id"])
        print(f"  ℹ Cargo Coordenador já existe para dept 4 (id={novo_cargo_id})")
    else:
        novo_cargo_id = max_id(ws_cargos) + 1
        ws_cargos.append([novo_cargo_id, "Coordenador", 2, None, 4])
        print(f"  ✓ Cargo Coordenador criado: id={novo_cargo_id}, dept=4, nivel=2, peso=None")

    # ── 3. Adiciona Patrice como Coordenadora de Seguro Internacional ───────
    print("\n[3] Adicionando Patrice como Coordenadora de Seguro Internacional (dept 4)")
    _, colabs = sheet_to_dicts(ws_colab)

    # Encontra Patrice no dept 1
    patrice = next((c for c in colabs if str(c.get("id", "")).strip() == "2"), None)
    if not patrice:
        print("  ✗ Patrice (id=2) não encontrada!")
    else:
        # Verifica se já tem entrada no dept 4
        patrice_dept4 = next(
            (c for c in colabs
             if normalizar_nome(str(c.get("nome", ""))) == normalizar_nome(str(patrice.get("nome", "")))
             and str(c.get("departamento_id", "")).strip() == "4"),
            None
        )
        if patrice_dept4:
            print(f"  ℹ Patrice já tem entrada em dept 4 (id={patrice_dept4['id']})")
        else:
            novo_colab_id = max_id(ws_colab) + 1
            # Detecta colunas da sheet colaboradores
            headers_colab = [str(ws_colab.cell(1, c).value or "") for c in range(1, ws_colab.max_column + 1)]
            nova_linha = []
            for h in headers_colab:
                h_low = h.lower().strip()
                if h_low == "id":
                    nova_linha.append(novo_colab_id)
                elif h_low == "nome":
                    nova_linha.append(patrice.get("nome"))
                elif h_low == "cargo_id":
                    nova_linha.append(novo_cargo_id)
                elif h_low == "departamento_id":
                    nova_linha.append(4)
                elif h_low == "empresa":
                    nova_linha.append(patrice.get("empresa"))
                elif h_low == "cidade":
                    nova_linha.append(patrice.get("cidade"))
                elif h_low == "tipo_contrato":
                    nova_linha.append(patrice.get("tipo_contrato"))
                elif h_low == "gestor_direto":
                    nova_linha.append(patrice.get("gestor_direto"))
                elif h_low == "data_entrada":
                    nova_linha.append(patrice.get("data_entrada"))
                elif h_low == "data_saida":
                    nova_linha.append(None)
                elif h_low == "status":
                    nova_linha.append("Ativo")
                elif h_low == "responsavel_direto":
                    nova_linha.append(None)
                else:
                    nova_linha.append(None)
            ws_colab.append(nova_linha)
            print(f"  ✓ Patrice adicionada em dept 4: id={novo_colab_id}, cargo_id={novo_cargo_id}")

    # ── 4. Ana Luísa (id=19): cargo_id 7 → 6 (Assistente → Analista Júnior)
    print("\n[4] Atualizando Ana Luísa (id=19): cargo_id 7 → 6")
    row_ana = find_row(ws_colab, col_idx(ws_colab, "id"), 19)
    if row_ana:
        col_cargo = col_idx(ws_colab, "cargo_id")
        ws_colab.cell(row=row_ana, column=col_cargo, value=6)
        print(f"  ✓ Ana Luísa linha {row_ana}: cargo_id → 6 (Analista Júnior)")
    else:
        print("  ✗ Ana Luísa (id=19) não encontrada!")

    # ── 5. Beatriz Vieira (id=49): cargo_id 7 → 8 (Assistente → Estagiário)
    print("\n[5] Atualizando Beatriz Vieira (id=49): cargo_id 7 → 8")
    row_bea = find_row(ws_colab, col_idx(ws_colab, "id"), 49)
    if row_bea:
        col_cargo = col_idx(ws_colab, "cargo_id")
        ws_colab.cell(row=row_bea, column=col_cargo, value=8)
        print(f"  ✓ Beatriz Vieira linha {row_bea}: cargo_id → 8 (Estagiária)")
    else:
        print("  ✗ Beatriz Vieira (id=49) não encontrada!")

    # ── 6. Vitória (id=23): desligamento 2026-04-13 ──────────────────────
    print("\n[6] Desligamento Vitória (id=23) em 2026-04-13")
    row_vit = find_row(ws_colab, col_idx(ws_colab, "id"), 23)
    if row_vit:
        col_saida = col_idx(ws_colab, "data_saida")
        col_status = col_idx(ws_colab, "status")
        ws_colab.cell(row=row_vit, column=col_saida, value=date(2026, 4, 13))
        ws_colab.cell(row=row_vit, column=col_status, value="Inativo")

        # Busca nome e cargo da Vitória para o histórico
        col_nome_c = col_idx(ws_colab, "nome")
        col_cargo_c = col_idx(ws_colab, "cargo_id")
        nome_vit = ws_colab.cell(row=row_vit, column=col_nome_c).value
        cargo_id_vit = ws_colab.cell(row=row_vit, column=col_cargo_c).value

        # Pega nome do cargo
        _, cargos_atuais = sheet_to_dicts(ws_cargos)
        cargo_obj = next(
            (c for c in cargos_atuais if str(c.get("id", "")).strip() == str(cargo_id_vit).strip()),
            None
        )
        cargo_nome_vit = cargo_obj["nome"] if cargo_obj else str(cargo_id_vit)

        # Adiciona entrada no Histórico
        novo_hist_id = max_id(ws_hist) + 1
        ws_hist.append([
            novo_hist_id,
            23,
            nome_vit,
            "Desligamento",
            cargo_nome_vit,
            date(2026, 4, 13),
            "Desligamento registrado em 13/04/2026",
        ])
        print(f"  ✓ Vitória linha {row_vit}: data_saida=2026-04-13, status=Inativo")
        print(f"  ✓ Histórico id={novo_hist_id}: Desligamento de {nome_vit}")
    else:
        print("  ✗ Vitória (id=23) não encontrada!")

    # ── 7. Juliano (id=20): responsavel_direto = Patrice ─────────────────
    print("\n[7] Atualizando Juliano (id=20): responsavel_direto = Patrice Zeidler Basso")
    row_jul = find_row(ws_colab, col_idx(ws_colab, "id"), 20)
    if row_jul:
        col_resp = col_idx(ws_colab, "responsavel_direto")
        if col_resp:
            ws_colab.cell(row=row_jul, column=col_resp, value="Patrice Zeidler Basso")
            print(f"  ✓ Juliano linha {row_jul}: responsavel_direto → Patrice Zeidler Basso")
        else:
            print("  ✗ Coluna responsavel_direto não encontrada!")
    else:
        print("  ✗ Juliano (id=20) não encontrado!")

    # ── 8. Salva arquivo ──────────────────────────────────────────────────
    print("\n[8] Salvando dados.xlsx...")
    wb.save(DATA_PATH)
    print("  ✓ Arquivo salvo com sucesso!")

    # ── 9. Recalcula Manpower para Abril 2026 ─────────────────────────────
    print("\n[9] Recalculando Manpower de Abril/2026...")

    # Recarrega com dados atualizados
    wb2 = load_workbook(DATA_PATH)
    ws_colab2 = wb2["Colaboradores"]
    ws_cargos2 = wb2["Cargos"]
    ws_deptos2 = wb2["Departamentos"]
    ws_mp2 = wb2["ManpowerMensal"]
    ws_perf2 = wb2["Performance"]

    _, colabs2 = sheet_to_dicts(ws_colab2)
    _, cargos2 = sheet_to_dicts(ws_cargos2)
    _, deptos2 = sheet_to_dicts(ws_deptos2)

    cargos_map = {str(c["id"]): c for c in cargos2}
    ano, mes = 2026, 4
    referencia = date(ano, mes, monthrange(ano, mes)[1])

    manpower_dept = {str(d["id"]): 0.0 for d in deptos2}
    for c in colabs2:
        if status_efetivo(c, referencia) != "Ativo":
            continue
        cargo = cargos_map.get(str(c.get("cargo_id", "")))
        peso = cargo.get("peso_manpower") if cargo else None
        if peso is not None:
            try:
                dept_id = str(c.get("departamento_id", ""))
                if dept_id in manpower_dept:
                    manpower_dept[dept_id] += float(peso)
            except (TypeError, ValueError):
                pass

    print(f"\n  Manpower por departamento em {referencia}:")
    for dept in deptos2:
        dept_id = str(dept["id"])
        mp = manpower_dept.get(dept_id, 0.0)
        print(f"    Dept {dept['id']} ({dept['nome']}): {round(mp, 2)}")

    # Upsert ManpowerMensal
    for dept in deptos2:
        dept_id_int = dept["id"]
        dept_id_str = str(dept_id_int)
        mp_total = round(manpower_dept.get(dept_id_str, 0.0), 2)

        row_found = None
        for row in ws_mp2.iter_rows(min_row=2):
            r_ano = row[0].value
            r_mes = row[1].value
            r_dept = row[2].value
            if r_ano == ano and r_mes == mes and (
                str(r_dept).strip() == dept_id_str or r_dept == dept_id_int
            ):
                row_found = row[0].row
                break

        if row_found:
            ws_mp2.cell(row=row_found, column=4, value=mp_total)
            print(f"    → ManpowerMensal atualizado: dept {dept_id_int} = {mp_total}")
        else:
            ws_mp2.append([ano, mes, dept_id_int, mp_total])
            print(f"    → ManpowerMensal inserido: dept {dept_id_int} = {mp_total}")

    # Atualiza Performance (apenas Importação, dept 1)
    importacao_ids = [
        str(d["id"]) for d in deptos2
        if str(d.get("nome", "")).strip().lower() == "importação"
    ]
    if not importacao_ids:
        importacao_ids = [str(d["id"]) for d in deptos2]

    manpower_perf = sum(
        manpower_dept.get(did, 0.0) for did in importacao_ids
    )
    manpower_perf = round(manpower_perf, 2)
    print(f"\n  Manpower de Performance (Importação): {manpower_perf}")

    for row in ws_perf2.iter_rows(min_row=2):
        if row[0].value == ano and row[1].value == mes:
            volume_score = row[2].value or 0
            meta = row[5].value
            performance = round(float(volume_score) / manpower_perf, 2) if manpower_perf and volume_score else 0
            pct_meta = round(performance / float(meta), 4) if meta and float(meta) > 0 else None
            ws_perf2.cell(row=row[0].row, column=4, value=manpower_perf)
            ws_perf2.cell(row=row[0].row, column=5, value=performance)
            ws_perf2.cell(row=row[0].row, column=7, value=pct_meta)
            print(f"  ✓ Performance abril/2026: manpower={manpower_perf}, perf={performance}")
            break
    else:
        print("  ℹ Sem registro de Performance para abril/2026 (sem volume_score cadastrado)")

    wb2.save(DATA_PATH)
    print("\n  ✓ Manpower salvo com sucesso!")
    print("\n✅ Todas as atualizações concluídas!")


def normalizar_nome(valor):
    """Normaliza nome removendo acentos e espaços extras."""
    import unicodedata, re
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


if __name__ == "__main__":
    main()
