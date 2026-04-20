#!/usr/bin/env python3
"""
Adiciona Beatriz Ferraz como nova colaboradora:
- Nome: Beatriz Ferraz
- Cargo: Assistente (cargo_id=7, peso_manpower=0.5)
- Departamento: Importação (departamento_id=1)
- Cidade/Empresa: Itajaí
- Gestor direto: Ana Luiza Silva (id=19)
- Responsável direto: Ana Luiza Silva
- Data de entrada: 2026-04-20

Efeitos em cascata:
1. Colaboradores: nova linha (id=63)
2. Historico: nova entrada (id=28, tipo=Entrada)
3. ManpowerMensal: Abril/2026, dept 1: 24.05 → 24.55 (+0.5)
"""
import os
import sys
from datetime import date

_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, _PROJECT_ROOT)

from openpyxl import load_workbook

DATA_PATH = os.path.join(_PROJECT_ROOT, "data", "dados.xlsx")


def col_idx(ws, header_name):
    for cell in ws[1]:
        if str(cell.value or "").strip().lower() == header_name.lower():
            return cell.column
    return None


def max_id(ws):
    max_val = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            try:
                max_val = max(max_val, int(row[0]))
            except (ValueError, TypeError):
                pass
    return max_val


def main():
    print(f"Carregando {DATA_PATH}...")
    wb = load_workbook(DATA_PATH)

    ws_colab = wb["Colaboradores"]
    ws_hist = wb["Historico"]
    ws_mp = wb["ManpowerMensal"]

    # ── 1. Adiciona Beatriz Ferraz em Colaboradores ───────────────────────
    print("\n[1] Adicionando Beatriz Ferraz em Colaboradores")

    novo_id = max_id(ws_colab) + 1
    data_entrada = date(2026, 4, 20)

    headers = [str(ws_colab.cell(1, c).value or "") for c in range(1, ws_colab.max_column + 1)]
    nova_linha = []
    for h in headers:
        h_low = h.lower().strip()
        if h_low == "id":
            nova_linha.append(novo_id)
        elif h_low == "nome":
            nova_linha.append("Beatriz Ferraz")
        elif h_low == "cargo_id":
            nova_linha.append(7)
        elif h_low == "departamento_id":
            nova_linha.append(1)
        elif h_low == "empresa":
            nova_linha.append("Itajaí")
        elif h_low == "cidade":
            nova_linha.append("Itajaí")
        elif h_low == "tipo_contrato":
            nova_linha.append(None)
        elif h_low == "gestor_direto":
            nova_linha.append("Ana Luiza Silva")
        elif h_low == "data_entrada":
            nova_linha.append(data_entrada)
        elif h_low == "data_saida":
            nova_linha.append(None)
        elif h_low == "status":
            nova_linha.append("Ativo")
        elif h_low == "responsavel_direto":
            nova_linha.append("Ana Luiza Silva")
        else:
            nova_linha.append(None)

    ws_colab.append(nova_linha)
    print(f"  ✓ Beatriz Ferraz inserida: id={novo_id}, cargo_id=7 (Assistente), dept=1 (Importação), cidade=Itajaí")
    print(f"    gestor_direto=Ana Luiza Silva, responsavel_direto=Ana Luiza Silva, data_entrada=2026-04-20")

    # ── 2. Adiciona entrada no Histórico ──────────────────────────────────
    print("\n[2] Registrando entrada no Histórico")

    novo_hist_id = max_id(ws_hist) + 1
    ws_hist.append([
        novo_hist_id,
        novo_id,
        "Beatriz Ferraz",
        "Entrada",
        "Assistente",
        data_entrada,
        "Contratação registrada em 20/04/2026",
    ])
    print(f"  ✓ Histórico id={novo_hist_id}: Entrada de Beatriz Ferraz em 20/04/2026")

    # ── 3. Salva o arquivo ────────────────────────────────────────────────
    print("\n[3] Salvando dados.xlsx...")
    wb.save(DATA_PATH)
    print("  ✓ Arquivo salvo com sucesso!")

    # ── 4. Atualiza ManpowerMensal Abril/2026 dept 1 ─────────────────────
    print("\n[4] Atualizando ManpowerMensal: Abril/2026, Importação (dept 1)")
    print("    Assistente peso_manpower=0.5 → 24.05 + 0.50 = 24.55")

    wb2 = load_workbook(DATA_PATH)
    ws_mp2 = wb2["ManpowerMensal"]

    row_found = None
    for row in ws_mp2.iter_rows(min_row=2):
        if row[0].value == 2026 and row[1].value == 4 and row[2].value == 1:
            row_found = row[0].row
            break

    if row_found:
        ws_mp2.cell(row=row_found, column=4, value=24.55)
        print(f"  ✓ ManpowerMensal linha {row_found}: dept 1, abr/2026 → 24.55")
    else:
        ws_mp2.append([2026, 4, 1, 24.55])
        print("  ✓ ManpowerMensal: nova linha inserida (2026, 4, 1, 24.55)")

    wb2.save(DATA_PATH)
    print("  ✓ ManpowerMensal salvo com sucesso!")

    print("\n✅ Todas as atualizações concluídas com sucesso!")
    print("\n── RESUMO ──────────────────────────────────────────────────────────")
    print(f"  Colaboradores: Beatriz Ferraz adicionada (id={novo_id})")
    print(f"  Historico: Entrada registrada (id={novo_hist_id})")
    print(f"  ManpowerMensal: Importação Abr/2026 → 24.55 (era 24.05, +0.50)")
    print(f"  Performance: sem registro de Abr/2026 → nenhuma alteração necessária")


if __name__ == "__main__":
    main()
