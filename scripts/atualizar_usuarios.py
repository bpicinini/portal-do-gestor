#!/usr/bin/env python3
"""Cria usuário Mariana Strick e atualiza senha da Patrice Basso diretamente no dados.xlsx."""
import hashlib
import os
import sys
from datetime import datetime

from openpyxl import load_workbook

DATA_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "dados.xlsx")
SHEET = "Usuarios"
HEADERS = ["id", "nome", "email", "perfil", "status", "modulos", "senha_hash", "senha_salt", "ultimo_login", "criado_em"]


def _hash(senha: str, salt_hex: str | None = None):
    salt = bytes.fromhex(salt_hex) if salt_hex else os.urandom(16)
    digest = hashlib.pbkdf2_hmac("sha256", senha.encode("utf-8"), salt, 240000)
    return digest.hex(), salt.hex()


def _proximo_id(ws):
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            try:
                max_id = max(max_id, int(row[0]))
            except (TypeError, ValueError):
                pass
    return max_id + 1


def criar_usuario(ws, nome, email, perfil, senha):
    email = email.strip().lower()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[2] or "").strip().lower() == email:
            print(f"  Usuário '{email}' já existe, pulando.")
            return
    senha_hash, senha_salt = _hash(senha)
    modulos = "Home|Organograma|KPIs|Controle de Pessoas"
    ws.append([
        _proximo_id(ws), nome, email, perfil, "Ativo",
        modulos, senha_hash, senha_salt, None, datetime.now()
    ])
    print(f"  Criado: {nome} <{email}>")


def atualizar_senha(ws, email, nova_senha):
    email = email.strip().lower()
    for row in ws.iter_rows(min_row=2):
        if str(row[2].value or "").strip().lower() == email:
            senha_hash, senha_salt = _hash(nova_senha)
            row[6].value = senha_hash  # col 7 (index 6) = senha_hash
            row[7].value = senha_salt  # col 8 (index 7) = senha_salt
            print(f"  Senha atualizada: {email}")
            return
    print(f"  AVISO: usuário '{email}' não encontrado.")


wb = load_workbook(DATA_PATH)
ws = wb[SHEET]

print("=== Criando Mariana Strick ===")
criar_usuario(ws, "Mariana Strick", "mariana.strick@3scorporate.com", "Usuário", "mari@3s2026")

print("=== Atualizando senha da Patrice Basso ===")
atualizar_senha(ws, "patrice.basso@3scorporate.com", "patrice@3s2026")

wb.save(DATA_PATH)
print("=== dados.xlsx salvo com sucesso. ===")
