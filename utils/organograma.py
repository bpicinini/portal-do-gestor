import os
import re
import unicodedata
from collections import defaultdict
from functools import lru_cache

from openpyxl import load_workbook


_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_MEGAZORD_PATH = os.path.join(_PROJECT_ROOT, "arquivos base", "PROJETO MEGAZORD.xlsx")


def _nivel(pessoa):
    try:
        return float(pessoa.get("cargo_nivel", 99) or 99)
    except (TypeError, ValueError):
        return 99.0


def _unidade_meta(pessoa):
    unidade = str(pessoa.get("unidade") or pessoa.get("empresa") or "").strip()
    cidade = str(pessoa.get("cidade") or "").strip()
    return unidade or cidade


def normalizar_nome(valor):
    texto = str(valor or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def _chaves_nome(valor):
    base = normalizar_nome(valor)
    if not base:
        return set()
    partes = base.split()
    chaves = {base}
    if partes:
        chaves.add(partes[0])
        chaves.add(partes[-1])
    if len(partes) >= 2:
        chaves.add(f"{partes[0]} {partes[-1]}")
    return {item for item in chaves if item}


def _pontuar_nome(candidato, referencia):
    cand = normalizar_nome(candidato)
    ref = normalizar_nome(referencia)
    if not cand or not ref:
        return -1
    if cand == ref:
        return 100
    if ref in cand or cand in ref:
        return 80
    cand_tokens = set(cand.split())
    ref_tokens = set(ref.split())
    inter = len(cand_tokens & ref_tokens)
    if inter:
        return inter * 10
    return -1


def _buscar_melhor_pessoa(candidatos, referencia):
    melhor = None
    melhor_score = -1
    for pessoa in candidatos:
        score = _pontuar_nome(pessoa.get("nome"), referencia)
        if score > melhor_score:
            melhor = pessoa
            melhor_score = score
    return melhor if melhor_score >= 10 else None


@lru_cache(maxsize=1)
def carregar_sementes_megazord():
    """Lê a planilha-base para obter os reportes já explicitados no organograma."""
    sementes = defaultdict(list)
    if not os.path.exists(_MEGAZORD_PATH):
        return {}

    wb = load_workbook(_MEGAZORD_PATH, data_only=True, read_only=True)
    try:
        if "Estrutura Gabriel" in wb.sheetnames:
            ws = wb["Estrutura Gabriel"]
            # Mapeamentos explícitos visíveis na seção de assistentes.
            for linha in range(21, 27):
                subordinado = ws[f"C{linha}"].value
                analista = ws[f"D{linha}"].value
                if subordinado and analista:
                    sementes["Importação"].append((str(subordinado), str(analista)))
    finally:
        wb.close()

    return dict(sementes)


def construir_estrutura_reportes(colaboradores):
    por_departamento = defaultdict(list)
    for pessoa in colaboradores:
        por_departamento[pessoa["departamento_nome"]].append(pessoa)

    estrutura = []
    for departamento, pessoas in sorted(por_departamento.items()):
        pessoas_ordenadas = sorted(pessoas, key=lambda p: (_nivel(p), normalizar_nome(p["nome"])))
        lideres = [p for p in pessoas_ordenadas if 1 < _nivel(p) <= 2.5]
        analistas = [p for p in pessoas_ordenadas if 3 <= _nivel(p) < 7]
        base = [p for p in pessoas_ordenadas if _nivel(p) >= 7]

        # Grupos analista → subordinados (via responsavel_direto)
        grupos = {analista["id"]: [] for analista in analistas}
        # Grupos lider → subordinados diretos (base que aponta para lider)
        grupos_lideres_diretos = {lider["id"]: [] for lider in lideres}
        usados = set()

        for subordinado in base:
            resp = str(subordinado.get("responsavel_direto") or "").strip()
            if not resp:
                continue
            analista = _buscar_melhor_pessoa(analistas, resp)
            if analista:
                grupos[analista["id"]].append(subordinado)
                usados.add(subordinado["id"])
                continue
            lider = _buscar_melhor_pessoa(lideres, resp)
            if lider:
                grupos_lideres_diretos[lider["id"]].append(subordinado)
                usados.add(subordinado["id"])

        # Agrupa analistas sob seu lider via gestor_direto
        analistas_por_lider = {lider["id"]: [] for lider in lideres}
        analistas_sem_lider = []
        for analista in analistas:
            gestor = str(analista.get("gestor_direto") or "").strip()
            lider = _buscar_melhor_pessoa(lideres, gestor) if gestor else None
            if lider:
                analistas_por_lider[lider["id"]].append(analista)
            else:
                analistas_sem_lider.append(analista)

        def _grupos_analista(lista):
            return [
                {
                    "analista": a,
                    "reportes": sorted(
                        grupos.get(a["id"], []),
                        key=lambda p: (_nivel(p), normalizar_nome(p["nome"])),
                    ),
                }
                for a in lista
            ]

        secoes_lider = [
            {
                "lider": lider,
                "grupos_analistas": _grupos_analista(analistas_por_lider.get(lider["id"], [])),
                "reportes_diretos": sorted(
                    grupos_lideres_diretos.get(lider["id"], []),
                    key=lambda p: (_nivel(p), normalizar_nome(p["nome"])),
                ),
            }
            for lider in lideres
        ]

        sem_alocacao = [p for p in base if p["id"] not in usados]

        estrutura.append(
            {
                "departamento": departamento,
                "total_pessoas": len(pessoas),
                "lideres": lideres,
                "secoes_lider": secoes_lider,
                "grupos_sem_lider": _grupos_analista(analistas_sem_lider),
                "sem_alocacao": sorted(sem_alocacao, key=lambda p: normalizar_nome(p["nome"])),
            }
        )

    estrutura.sort(key=lambda b: -b["total_pessoas"])
    return estrutura
