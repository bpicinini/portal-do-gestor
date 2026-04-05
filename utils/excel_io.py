"""
Camada de persistência — lê e salva dados.xlsx via GitHub API (deploy)
ou arquivo local (desenvolvimento).
"""
import io
import os
import json
import base64
import streamlit as st
from openpyxl import Workbook, load_workbook

# Constantes das sheets
SHEET_DEPARTAMENTOS = "Departamentos"
SHEET_CARGOS = "Cargos"
SHEET_COLABORADORES = "Colaboradores"
SHEET_HISTORICO = "Historico"
SHEET_MANPOWER_MENSAL = "ManpowerMensal"
SHEET_PERFORMANCE = "Performance"
SHEET_USUARIOS = "Usuarios"

# Headers de cada sheet
HEADERS = {
    SHEET_DEPARTAMENTOS: ["id", "nome"],
    SHEET_CARGOS: ["id", "nome", "nivel", "peso_manpower", "departamento_id"],
    SHEET_COLABORADORES: [
        "id", "nome", "cargo_id", "departamento_id", "empresa", "cidade",
        "tipo_contrato", "gestor_direto", "data_entrada", "data_saida", "status"
    ],
    SHEET_HISTORICO: ["id", "colaborador_id", "nome", "tipo", "cargo", "data", "observacao"],
    SHEET_MANPOWER_MENSAL: ["ano", "mes", "departamento_id", "manpower_total"],
    SHEET_PERFORMANCE: ["ano", "mes", "volume_score", "manpower", "performance", "meta", "pct_meta"],
    SHEET_USUARIOS: [
        "id",
        "nome",
        "email",
        "perfil",
        "status",
        "modulos",
        "senha_hash",
        "senha_salt",
        "ultimo_login",
        "criado_em",
    ],
}

# Raiz do projeto — resolve tanto via __file__ quanto via CWD
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Fallback: se __file__ não resolve para a raiz correta, usar CWD
if not os.path.exists(os.path.join(_PROJECT_ROOT, "config", "settings.json")):
    _PROJECT_ROOT = os.getcwd()

_CONFIG_PATH = os.path.join(_PROJECT_ROOT, "config", "settings.json")
with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
    _SETTINGS = json.load(f)

_DATA_PATH = os.path.join(_PROJECT_ROOT, _SETTINGS["data_file"])


def _usar_github():
    """Verifica se deve usar GitHub como storage."""
    try:
        return bool(st.secrets.get("GITHUB_TOKEN") and st.secrets.get("GITHUB_REPO"))
    except Exception:
        return False


def _github_get_file():
    """Lê dados.xlsx do GitHub, retorna (bytes, sha)."""
    from github import Github
    g = Github(st.secrets["GITHUB_TOKEN"])
    repo = g.get_repo(st.secrets["GITHUB_REPO"])
    path = st.secrets.get("GITHUB_DATA_PATH", "data/dados.xlsx")
    try:
        contents = repo.get_contents(path)
        data = base64.b64decode(contents.content)
        return data, contents.sha
    except Exception:
        return None, None


def _github_put_file(file_bytes, sha=None):
    """Salva dados.xlsx no GitHub."""
    from github import Github
    g = Github(st.secrets["GITHUB_TOKEN"])
    repo = g.get_repo(st.secrets["GITHUB_REPO"])
    path = st.secrets.get("GITHUB_DATA_PATH", "data/dados.xlsx")
    encoded = base64.b64encode(file_bytes).decode()
    if sha:
        repo.update_file(path, "Atualização dados.xlsx", encoded, sha)
    else:
        repo.create_file(path, "Criação dados.xlsx", encoded)


def _garantir_sheets(wb):
    """Cria sheets faltantes e seus headers."""
    for sheet_name, headers in HEADERS.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
    # Remove sheet padrão vazia se existir
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]


def carregar_workbook():
    """Carrega o workbook (GitHub ou local), garante estrutura."""
    if _usar_github():
        data, sha = _github_get_file()
        if data:
            wb = load_workbook(io.BytesIO(data))
            wb._github_sha = sha
        else:
            wb = Workbook()
            wb._github_sha = None
    else:
        if os.path.exists(_DATA_PATH):
            wb = load_workbook(_DATA_PATH)
        else:
            os.makedirs(os.path.dirname(_DATA_PATH), exist_ok=True)
            wb = Workbook()
        wb._github_sha = None

    _garantir_sheets(wb)
    return wb


def salvar_workbook(wb):
    """Salva o workbook (GitHub ou local)."""
    if _usar_github():
        buffer = io.BytesIO()
        wb.save(buffer)
        _github_put_file(buffer.getvalue(), getattr(wb, '_github_sha', None))
    else:
        os.makedirs(os.path.dirname(_DATA_PATH), exist_ok=True)
        try:
            wb.save(_DATA_PATH)
        except PermissionError:
            raise IOError("Não foi possível salvar — o arquivo dados.xlsx está aberto no Excel. Feche e tente novamente.")


def ler_bytes_workbook():
    """Retorna bytes do workbook para download."""
    if _usar_github():
        data, _ = _github_get_file()
        if data:
            return data
    if os.path.exists(_DATA_PATH):
        with open(_DATA_PATH, "rb") as f:
            return f.read()
    return None


def proximo_id(ws):
    """Retorna o próximo id disponível para uma sheet."""
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None and isinstance(row[0], (int, float)):
            max_id = max(max_id, int(row[0]))
    return max_id + 1


def sheet_to_list(ws):
    """Converte sheet em lista de dicts (header na row 1)."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def encontrar_linha(ws, col_idx, valor):
    """Encontra o número da linha onde ws[row][col_idx] == valor. Retorna None se não achar."""
    _v = str(valor).strip() if valor is not None else None
    for row in ws.iter_rows(min_row=2):
        cell = row[col_idx - 1]
        _c = str(cell.value).strip() if cell.value is not None else None
        if _c == _v:
            return cell.row
    return None
